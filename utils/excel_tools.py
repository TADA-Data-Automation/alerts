from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd

def save_excel(summary_df, full_df, output_path):
    """Save summary and full data to Excel with formatting."""
    with pd.ExcelWriter(output_path, engine="openpyxl", mode="w") as writer:
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        full_df.to_excel(writer, sheet_name="details", index=False)

    apply_conditional_formatting(output_path)

def apply_conditional_formatting(output_path):
    """Apply color formatting to increase/decrease actions."""
    wb = load_workbook(output_path)
    ws = wb["summary"]

    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    action_col_idx = [cell.value for cell in ws[1]].index("action")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        action = row[action_col_idx].value
        if action == "increase":
            for cell in row:
                cell.fill = fill_green
        elif action == "decrease":
            for cell in row:
                cell.fill = fill_red

    wb.save(output_path)
